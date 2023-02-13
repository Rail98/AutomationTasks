#!/usr/bin/env python
# coding: utf-8

# #### Средний отчет
# В отчете будет несколько запросов. Используем библиотеку `threading`, чтобы запустить параллельно сразу несколько запросов.
# 
# Таблицы, которые нужны для отчетности, могут храниться на разных серверах. Для удобства, сделаем так, чтобы мы могли указывать, к какому серверу мы подключаемся.

# In[ ]:


print("Начало запуска Отчета №2...")


# In[1]:


# импортируем все нужные библиотеки для выполения кода
import pandas as pd
import pyodbc
import openpyxl as ox
from datetime import date
import calendar
import os
import time
from openpyxl.styles import Alignment # определяет положение значения в ячейки (слева, справа, по середине)
from openpyxl.styles import Font # может менять шрифт
from openpyxl.styles.borders import Border, Side # для добавления границ
from openpyxl.styles import PatternFill # для закрашивания
from openpyxl.styles import colors

#import threading
from threading import Thread

#pd.set_option("display.max_columns", 50)


# In[2]:


# Для конвертации расширения .ipynb в .py прописываем в командной строке:
# jupyter nbconvert Отчет_1.ipynb --to script


# In[4]:


# создаем список имен месяцов, чтобы затем указать имя текущего месяца на русском
name_of_months = ['Не указан', 'Январь', 'Февраль', 
                  'Март', 'Апрель', 'Май', 
                  'Июнь', 'Июль', 'Август', 
                  'Сентябрь', 'Октябрь', 'Ноябрь', 
                  'Декабрь']
current_month = int(date.today().strftime("%m")) # текущий месяц
current_date = date.today() # текущая дата


# In[9]:


# создаем папки, если её не существует
path_folders = os.path.join(current_date.strftime("%Y"), 
                            name_of_months[current_month], 
                            current_date.strftime("%Y.%m.%d")) # путь, состоящий только из папок
if not os.path.isdir(path_folders):
     os.makedirs(path_folders)


# #### Выгрузка данных

# Подключимся к базе данных.
# В начале коннектимся, а в конце закрываем соединение. В середине будем использовать контекстные менеджеры `with as` для каждого запроса.
# 

# In[ ]:


# создадим функцию для соединения с сервером
def connection(params):
    conn = pyodbc.connect(server=params['server_name'],
                          database=params['database_name'],
                          driver=params['driver_name'],
                          Trusted_Connection='yes')
    return conn

def read_sql(connection_params, path_to_query, result_list, index):
    conn = connection(connection_params)
    with open(path_to_query) as query:
        result_list[index] = pd.read_sql(query.read(), conn)
    conn.close()


# In[ ]:


print("Выполняется запрос 'Изменилась дата отгрузки.sql'...")


# In[ ]:


results = [None]*2


# In[ ]:


connection_params = {'server_name': 'DESKTOP-TI9PUNE\TESTSERVER',
                     'database_name': 'testdb'
                     'driver_name': '{ODBC Driver 17 for SQL Server}'} # указываем параметры соединения с серверов

path_to_query = os.path.join('SQL запросы', '(Тек.м) Акции по LKA с тм збк new.sql') # путь, где лежит запрос
Thread_1 = Thread(target=read_sql, args=(connection_params, path_to_query, results, 0)) # создаем объект Thread
Thread_1.start() # паралелльно запускаем запрос


connection_params = {'server_name': 'DESKTOP-TI9PUNE\TESTSERVER',
                     'database_name': 'testdb'
                     'driver_name': '{ODBC Driver 17 for SQL Server}'}
path_to_query = os.path.join('SQL запросы', '(Тек.м) ДОЧЕРНИЕ Акции по LKA с тм збк new.sql')
Thread_2 = Thread(target=read_sql, args=(connection_params, path_to_query, results, 0))
Thread_2.start()


# In[ ]:


# посмотрим, что получилось
Thread_2.join() # ждем, пока выполнится последний запрос
results


# In[ ]:


df_sheet1 = results[0]
df_sheet2 = results[1]

del results


# In[10]:


print('Запрос выполнен...')


# #### Обработка данных

# Далее можно провести обработку таблицы, используя библиотеку pandas. Например, избавиться от лишних пробелов в значениях, найти новые переменные (среднее, сумму и т.д.).   

# In[ ]:


# создадим функцию, которая будет определять название месяца по его номеру
def replace_func(row):
    return name_of_months[row]


# In[ ]:


# преобразуем входные данные
df_sheet1['month_name'] = df_sheet1['month_name'].str.strip() # избавляемся от пустых строк в начале и в конце строки
df_sheet2['month_name'] = df_sheet2['month_name'].str.strip()

df_sheet1.loc[df_sheet1['month_name'] != '', 'month_name'] = pd.to_datetime(df_sheet1[df_sheet1['month_name'] != '']['month_name']).dt.month.astype(int) # присваиваем не пустым значения их номер месяца
df_sheet2.loc[df_sheet2['month_name'] != '', 'month_name'] = pd.to_datetime(df_sheet2[df_sheet2['month_name'] != '']['month_name']).dt.month.astype(int)

df_sheet1.loc[df_sheet1['month_name'] == '', 'month_name'] = 0 # пустые ячейки заполняем нулями
df_sheet2.loc[df_sheet2['month_name'] == '', 'month_name'] = 0

#pd.to_datetime(df[df[''] != ''][''], format='%d.%m.%Y').dt.month


# In[ ]:


# применим функцию replace_func
df_sheet1['month_name'] = df_sheet1['month_name'].apply(replace_func)
df_sheet2['month_name'] = df_sheet2['month_name'].apply(replace_func)


# In[ ]:


# проводим чистку от лишних пробелов
list_columns = ['shipment_start', 'shipment_end', 'holding_start', 'holding_end']
for col in list_columns:
    df_sheet1[col] = df_sheet1[col].str.strip() 
    df_sheet2[col] = df_sheet2[col].str.strip() 


# В этом отчете нужно отобразить две дополнительные колонки в виде формул (а не просто значение).

# In[ ]:


# создадим колонки kod1 и kod2
df_sheet1['kod1'] = range(6, len(df_sheet1)+6)
df_sheet2['kod1'] = range(6, len(df_sheet2)+6)
df_sheet1['kod2'] = range(6, len(df_sheet1)+6)
df_sheet2['kod2'] = range(6, len(df_sheet2)+6)


# In[ ]:


# создаем функции для присвоение формул в этих колонках
def formula_kod1(row):
    return f'=CONCATENATE(A{row},"S",C{row},"S",G{row})'
def formula_kod2(row):
    return f'=CONCATENATE(A{row},"S",C{row},"d",SUM(P{row}:S{row}),"s",T{row},"A",M{row})'    


# In[ ]:


# применяем функции для kod1 и kod2
df_sheet1['kod1'] = df_sheet1['kod1'].apply(formula_kod1)
df_sheet2['kod1'] = df_sheet2['kod1'].apply(formula_kod1)
df_sheet1['kod2'] = df_sheet1['kod2'].apply(formula_kod2)
df_sheet2['kod2'] = df_sheet2['kod2'].apply(formula_kod2)


# #### Внесение данных в эксель

# Теперь нужно залить данные в эксель. Для этого мы заранее подготовили шаблон (шапку), в которую мы и зальем наши данные.

# In[11]:


# указываем путь, в котором хрантся наш шаблон, и путь, куда сохраним уже готовый отчет
path_from = os.path.join('Excel шаблоны',
                         'Акции по LKA (Текущий месяц) шаблон.xlsx')
path_to = os.path.join(path_folders, 
                       f'Акции по LKA ({name_of_months[current_month]}) {date.today().strftime("%Y.%m.%d")}.xlsx') 


# In[12]:


print('Создается эксель файл...')


# In[ ]:


get_ipython().run_cell_magic('time', '', '\nwb = ox.load_workbook(path_from) # загружаем наш шаблон\nws = wb[\'По головным\'] # выбираем первый лист\n\nws.cell(1, 2).value = f\'Реестр акций по локальным ключевым клиентам (головные сети) за {name_of_months[current_month]}\'\nws.cell(2, 2).value = f\'Дата: {current_date.strftime("%d.%m.%Y")}\'\n\nfor ir in range(df_sheet1.shape[0]):\n    for ic in range(df_sheet1.shape[1]):\n        ws.cell(6 + ir, 1 + ic).value = df_sheet1.iloc[ir][ic] # заносим значения по ячеечкам\n    \n    for ic in [16, 17, 18, 19]:\n        ws.cell(6 + ir, ic).number_format = \'ДД.ММ.ГГГГ\' # задаем формат ячеек\n        ws.cell(6 + ir, ic).alignment = Alignment(horizontal=\'right\') # задаем выравнивание (по правому краю)\n    \n    for ic in [28, 29, 30, 31, 32, 33, 34, 35]:\n        ws.cell(6 + ir, ic).number_format = \'#,##0\' # задаем формат ячеек\n\n        \nws = wb[\'По дочерним\'] # выбираем второй лист\n\nfor ir in range(df_sheet2.shape[0]):\n    for ic in range(df_sheet2.shape[1]):\n        ws.cell(6 + ir, 1 + ic).value = df_sheet2.iloc[ir][ic] # заносим значения по ячеечкам\n        \n    for ic in [16, 17, 18, 19]:\n        ws.cell(6 + ir, ic).number_format = \'ДД.ММ.ГГГГ\' # задаем формат ячеек\n        ws.cell(6 + ir, ic).alignment = Alignment(horizontal=\'right\') # задаем выравнивание (по правому краю)\n    \n    for ic in [28, 29, 30, 31, 32, 33, 34, 35]:\n        ws.cell(6 + ir, ic).number_format = \'#,##0\' # задаем формат ячеек\n        \n\n# ws.print_area = f"A1:AI{5+len(df)}" # определяем область печати (нужно указать вручную только диапозон столбцов, а строки сами подтянутся автоматом)\n\nwb.save(path_to) # сохраняемся')


# In[13]:


print(f''''Изменилась дата отгрузки от {date.today().strftime("%d.%m.%Y")}.xlsx' файл создан...''')


# In[ ]:


print('-----')
print('-----')
print('Успех! Отчет выполнен. Можно закрыть консоль.')
print('-----')
print('-----')
time.sleep(3600*23)


# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:




