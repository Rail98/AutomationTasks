#!/usr/bin/env python
# coding: utf-8

# #### Для каждого отчета потребуется:
# - python-скрипт с расшерением `.py`;
# - файл с sql-запросами (`.sql`, `.txt`);
# - excel-файл, в котором будет сохранена шапка будущего запроса.

# #### Простой отчет
# Здесь будет выполняться только один запрос. Поэтому нам не потребуется `threading` (распараллеливание, многопоточность).

# In[ ]:


print("Начало запуска Отчета №1...")


# In[1]:


# импортируем все нужные библиотеки для выполения кода
import pandas as pd
import pyodbc
import openpyxl as ox
from datetime import date
import calendar
import os
import time
from openpyxl.styles import Alignment # определяет положение значения в ячейки (слева, справа, по середине)
from openpyxl.styles import Font # может менять шрифт
from openpyxl.styles.borders import Border, Side # для добавления границ
from openpyxl.styles import PatternFill # для закрашивания
from openpyxl.styles import colors

#pd.set_option("display.max_columns", 50)


# In[2]:


# Для конвертации расширения .ipynb в .py прописываем в командной строке:
# jupyter nbconvert Изменилась_дата_отгрузки_python.ipynb --to script


# In[3]:


# указываем параметры соединения
sql_server = '192.168.2.128'
database = 'nefco'
driver = '{SQL Server Native Client 10.0}'
conn_string = f'SERVER={sql_server};DATABASE={database};DRIVER={driver};Trusted_connection=Yes'


# In[4]:


# создаем список имен месяцов, чтобы затем указать имя текущего месяца на русском
name_of_months = ['Не указан', 'Январь', 'Февраль', 
                  'Март', 'Апрель', 'Май', 
                  'Июнь', 'Июль', 'Август', 
                  'Сентябрь', 'Октябрь', 'Ноябрь', 
                  'Декабрь']
current_month = int(date.today().strftime("%m")) # текущий месяц
current_date = date.today() # текущая дата


# In[9]:


# создаем папки, если её не существует
path_folders = os.path.join(current_date.strftime("%Y"), 
                            name_of_months[current_month], 
                            current_date.strftime("%Y.%m.%d")) # путь, состоящий только из папок
if not os.path.isdir(path_folders):
     os.makedirs(path_folders)


# #### Выгрузка данных

# Подключимся к базе данных.
# В начале коннектимся, а в конце закрываем соединение. В середине будем использовать контекстные менеджеры `with as` для каждого запроса.
# 

# In[ ]:


conn = pyodbc.connect(conn_string) # подключаемся к бд


# In[ ]:


print("Выполняется запрос 'Изменилась дата отгрузки.sql'...")


# In[ ]:


get_ipython().run_cell_magic('time', '', "path_to_query = os.path.join('SQL запросы',\n                             'Изменилась дата отгрузки.sql') # путь, в котором хранится запрос\n# считываем запрос, и сохраяем его в датасет\nwith open(path_to_query) as query:\n    df = pd.read_sql(query.read(), conn)")


# In[ ]:


conn.close() # отключаемся от бд


# In[10]:


print('Запрос выполнен...')


# #### Обработка данных

# Далее можно провести обработку таблицы, используя библиотеку pandas. Например, избавиться от лишних пробелов в значениях, найти новые переменные (среднее, сумму и т.д.).   

# In[ ]:


# создадим новую колонку, в котором нумерация начинается с 1 
df.index += 1
df = df.reset_index()


# In[ ]:


# удалим пробелы в начале и в конце слов у некоторых столбцов
columns = ['date_imp', 'date_otgr_pl1', 'date_otgr_fact', 'date_otgr_pl2', 'pl_vchera', 'pl_pozavchera', 'pl_ish']
for col in columns:
    df[col] = df[col].str.strip()


# #### Внесение данных в эксель

# Теперь нужно залить данные в эксель. Для этого мы заранее подготовили шаблон (шапку), в которую мы и зальем наши данные.

# In[11]:


# указываем путь, в котором хрантся наш шаблон, и путь, куда сохраним уже готовый отчет
path_from = os.path.join('Excel шаблоны',
                         'Изменилась дата отгрузки шаблон.xlsx')
path_to = os.path.join(path_folders, 
                       'Изменилась дата отгрузки от', 
                       date.today().strftime("%d.%m.%Y"), 
                       '.xlsx') 


# In[12]:


print('Создается эксель файл...')


# In[ ]:


get_ipython().run_cell_magic('time', '', '\nwb = ox.load_workbook(path_from) # загружаем наш шаблон\nws = wb.active # выбираем лист \n\nws.cell(2, 3).value = f\'Дата отчета: {current_date.strftime("%d.%m.%Y")} (время выгрузки 08:00)\'\n\nws.cell(6, 15).value = f\'=O7+O8+SUMIFS(O10:O{9+len(df)},B10:B{9+len(df)},"=17")\'\nws.cell(7, 15).value = f\'=SUMIFS(O10:O{9+len(df)},B10:B{9+len(df)},"<>13",B10:B{9+len(df)},"<>17")\'\nws.cell(8, 15).value = f\'=SUMIFS(O10:O{9+len(df)},B10:B{9+len(df)},"=13")\'\nws.cell(9, 15).value = f\'=SUMIF(W10:W{9+len(df)},"=0",O10:O{9+len(df)})\'\nws.cell(9, 15).value = f\'=SUMIF(W10:W{9+len(df)},"=0",O10:O{9+len(df)})\'\n#ws.cell(9, 15).font = Font(bold=True) # можно задать жирный шрифт отсюда или указать в шаапке шаблона\n\nfor ir in range(df.shape[0]):\n    for ic in range(df.shape[1]):\n        ws.cell(10 + ir, 1 + ic).value = df.iloc[ir][ic] # заносим значения по ячеечкам\n    ws.cell(10 + ir, 8).number_format = \'dd mm yyyy\' # задаем формат ячеек\n    ws.cell(10 + ir, 8).alignment = Alignment(horizontal=\'right\') # задаем выравнивание (по правому краю)\n    ws.cell(10 + ir, 9).number_format = \'dd mm yyyy\'\n    ws.cell(10 + ir, 9).alignment = Alignment(horizontal=\'right\')\n    ws.cell(10 + ir, 15).number_format = \'000000\'\n    ws.cell(10 + ir, 16).number_format = \'dd mm yyyy\'\n    ws.cell(10 + ir, 16).alignment = Alignment(horizontal=\'right\')\n    ws.cell(10 + ir, 17).number_format = \'dd mm yyyy\'\n    ws.cell(10 + ir, 17).alignment = Alignment(horizontal=\'right\')\n    ws.cell(10 + ir, 18).number_format = \'dd mm yyyy\'\n    ws.cell(10 + ir, 18).alignment = Alignment(horizontal=\'right\')\n    ws.cell(10 + ir, 19).number_format = \'dd mm yyyy\'\n    ws.cell(10 + ir, 19).alignment = Alignment(horizontal=\'right\')\n    ws.cell(10 + ir, 22).value = f\'=MONTH($P{10+ir})=MONTH($S{10+ir})\'\n    ws.cell(10 + ir, 23).value = f\'=$V{10+ir}+0\'\n#print(ws.cell(11, 19).number_format)\n\nws.print_area = f"A1:V{9+len(df)}" # определяем область печати (нужно указать вручную только диапозон столбцов, а строки сами подтянутся автоматом)\n\nwb.save(path_to) # сохраняемся')


# In[13]:


print(f''''Изменилась дата отгрузки от {date.today().strftime("%d.%m.%Y")}.xlsx' файл создан...''')


# In[ ]:


print('-----')
print('-----')
print('Успех! Отчет выполнен. Можно закрыть консоль.')
print('-----')
print('-----')
time.sleep(3600*23)


# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:




